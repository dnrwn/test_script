import openpyxl,tkinter.filedialog, os, json, datetime, logging

test_time = datetime.datetime.now().strftime('%y%m%d_%H%M%S')

def console_logger():
    logging.basicConfig(filename=f'log/{test_time}.log/', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    return logging.getLogger()

# file_read, sheet_read 함수 정리할 필요가 있을지 검토

def file_read(a='1'):
    console_logger().debug(file_read.__name__)
    # a = input('1 : 자동 실행, 2 : 직접 선택\n')
    if a == '1':  # 동일 디렉토리에 있는 xlsx 파일 중 첫 번째 파일 선택
        file = os.listdir(os.getcwd())
        return openpyxl.load_workbook([s for s in file if '.xlsx' in s][0])
    elif a == '2':  # 탐색기로 파일 선택
        ################## TBD ##################
        return openpyxl.load_workbook(tkinter.filedialog.askopenfilename(initialdir='', title='한 개의 파일 선택'))
    else:
        return console_logger().info(file_read.__name__, '잘못된 선택')


def sheet_read(a='1'):
    console_logger().debug(sheet_read.__name__)
    if a == '1':  # 특정 sheet를 읽을 경우
        return 'TC'
    else:   # 여러 sheet를 읽을 경우
        ################## TBD ##################
        for sheet_nm in file_read().sheetnames:
            sheet = file_read()[sheet_nm]
        return None


class FUNC:
    def __init__(self):
        console_logger().debug(FUNC.__name__)
        self.file = file_read()
        self.file_sheet = self.file[sheet_read()]
        self.case_item = {}
        self.case_result = {}

    def test_item_array(self):
        console_logger().debug(FUNC.test_item_array.__name__)
        i = 1  # 열
        for col_data in self.file_sheet.iter_rows(min_col=1, min_row=2, max_col=self.file_sheet.max_column,
                                                  max_row=self.file_sheet.max_row):
            for item_v in col_data:
                if (i + 1) % self.file_sheet.max_column == 0 or i % self.file_sheet.max_column == 0:
                    globals()[f'em_{i}'] = item_v.coordinate
                else:
                    globals()[f'em_{i}'] = '' if item_v.value is None else item_v.value

                if i % self.file_sheet.max_column == 0:  # 행의 모든 열을 변수에 삽입 후 진입
                    z = i - (self.file_sheet.max_column - 1)  # 두 번째 행(row) 부터 읽기 때문에 -1 추가
                    # case_item 변수에 data 삽입
                    self.case_item[globals()[f"em_{z + 0}"]] = {
                        "function_name": globals()[f'em_{z + 1}'],
                        "number": globals()[f'em_{z + 2}'],
                        "idx": globals()[f'em_{z + 5}'],
                        "input_1": globals()[f'em_{z + 6}'],
                        "input_2": globals()[f'em_{z + 7}'],
                        "input_3": globals()[f'em_{z + 8}'],
                        "input_4": globals()[f'em_{z + 9}']
                    }
                    # case_result 변수에 data 삽입
                    self.case_result[globals()[f"em_{z + 0}"]] = {
                        "function_name": globals()[f'em_{z + 1}'],
                        "number": globals()[f'em_{z + 2}'],
                        "Expected_result": globals()[f'em_{z + 10}'],
                        "Actual_result": globals()[f'em_{z + 11}'],  # 결과 기록 위치 저장
                        "record": globals()[f'em_{z + 12}']  # 결과 기록 위치 저장
                    }
                i += 1
        # 변수에 삽입된 data 를 json 형태로 전달
        return json.dumps({
            "item": self.case_item,
            "result": self.case_result
        })

    def test_result_write(self, result_cell, record_cell, result, record=None):
        console_logger().debug(FUNC.test_result_write.__name__)
        self.file_sheet[result_cell] = result
        self.file_sheet[record_cell] = str(record) # Fail 일 경우 record 값이 list type 이므로 string type 로 변환 필요
        self.file.save(f'result/TC_result_{test_time}.xlsx')
        self.file.close()