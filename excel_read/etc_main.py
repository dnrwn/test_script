import openpyxl
import tkinter.filedialog
import datetime


def time():
    return datetime.datetime.now().strftime('%y%m%d_%H%M%S')


def value_setup():
    file = tkinter.filedialog.askopenfilename(initialdir='', title='한 개의 파일 선택', filetypes='')
    # askopenfilename 모듈이 튜플 형태로 return함
    load_wb = openpyxl.load_workbook(file)

    ##################
    # for sheet in load_wb.sheetnames: # 여러 시트를 읽어야 할 경우 필요한 부분
    #     sheet = load_wb[sheet]
    ##################

    sheet = load_wb['TC'] # TC 시트를 특정함
    for row_data in sheet.iter_rows(min_row=2, min_col=0): # 열 col, 행 row / max로 통제 필요한지 검토 필요
        i = 1
        for cell in row_data:
            # f'{}' 형태로 작성 시 괄호에 변수 삽입 가능하며 동적 변수 생성 가능
            globals()[f'cell_{i}'] = cell
            # 동적 변수 실행 완료 후 실제 함수 실행부, 마지막 열 까지 다 읽은 다음에 아래 코드 실행하기 위함
            if i == len(row_data):
                print(i)
                # 함수 실행 + 함수 인자에 parameter 대입 + result, return 값 비교
                if globals()[cell_2.value](cell_5.value) == cell_6.value:
                    # 대괄호 안의 value가 func 모듈의 함수명을 가르키고 있음. (모듈 호출 없이 바로 함수 호출하는 방식은 검토 필요)
                    # = a_1(9)
                    sheet[cell_7.coordinate] = 'PASS'

                    # cell_6 = cell_6.value.split(', ')
                    # print('PASS', cell_6.parent.title) # Sheet 이름
                    # print('PASS', cell_6.row) # Cell 행 위치
                    # print('PASS', cell_6.column) # Cell 열 위치
                else:
                    sheet[cell_7.coordinate] = 'FAIL'
            i += 1
        load_wb.save(f'TC_result_{datetime.time()}.xlsx')
        load_wb.close()
if __name__ == '__main__':
    value_setup()
